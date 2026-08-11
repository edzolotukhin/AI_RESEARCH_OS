"""P1-15.1 offline acceptance: XLSX content adapter (CASES 1–24)."""

from __future__ import annotations

import hashlib
import io
import time
import unittest
import zipfile
from datetime import date
from pathlib import Path
from unittest.mock import Mock, patch

from openpyxl import Workbook

from application.evidence.grounding import verify_grounding
from domain.sources.retrieval_status import RetrievalStatus
from domain.sources.source_candidate import SourceCandidate
from infrastructure.retrieval.http_source_retriever import HttpSourceRetriever
from infrastructure.retrieval.xlsx_text_extractor import (
    XLSX_CONTENT_TYPE,
    XlsxExtractionLimits,
    extract_xlsx_text,
)


REPO_ROOT = Path(__file__).resolve().parents[3]
DESNZ_PATH = (
    REPO_ROOT
    / "artifacts"
    / "acceptance"
    / "fixtures"
    / "desnz_heat_pump_deployment_2025_q2.xlsx"
)
DESNZ_SHA256 = "06ce4557f1766c53dc0cf42cc065ef32fec97e47ed921f7c0293addb7f1afed6"


def _xlsx_bytes(build) -> bytes:
    workbook = Workbook()
    build(workbook)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _mock_response(*, url: str, content_type: str, content: bytes, status: int = 200):
    response = Mock()
    response.url = url
    response.status_code = status
    response.headers = {"content-type": content_type}
    response.content = content
    return response


class P1151XlsxContentAdapterTests(unittest.TestCase):
    def test_case_01_simple_single_sheet_parses(self) -> None:
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws.title = "Data"
            ws["A1"] = "Period"
            ws["B1"] = "Installations"
            ws["A2"] = "2024"
            ws["B2"] = 43500

        result = extract_xlsx_text(_xlsx_bytes(build))
        self.assertTrue(result.ok)
        self.assertIn('[Sheet: Data]', result.text)
        self.assertIn("43500", result.text)
        self.assertEqual(result.metadata.get("parser"), "xlsx")
        self.assertFalse(result.metadata.get("workbook_truncated"))

    def test_case_02_multiple_visible_sheets_deterministic(self) -> None:
        def build(wb: Workbook) -> None:
            first = wb.active
            first.title = "Alpha"
            first["A1"] = "alpha-value"
            second = wb.create_sheet("Beta")
            second["A1"] = "beta-value"

        result = extract_xlsx_text(_xlsx_bytes(build))
        alpha_idx = result.text.index("[Sheet: Alpha]")
        beta_idx = result.text.index("[Sheet: Beta]")
        self.assertLess(alpha_idx, beta_idx)
        self.assertEqual(result.metadata.get("processed_sheets"), 2)

    def test_case_03_multirow_headers_preserve_context(self) -> None:
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws.title = "Geo"
            ws["A1"] = "Metric"
            ws["B1"] = "K04000001"
            ws["C1"] = "E92000001"
            ws["A2"] = "Installation quarter"
            ws["B2"] = "England and Wales"
            ws["C2"] = "England"
            ws["A3"] = "2024 Q2"
            ws["B3"] = 9470
            ws["C3"] = 7672

        result = extract_xlsx_text(_xlsx_bytes(build))
        self.assertTrue(result.ok)
        matching = [
            line
            for line in result.text.splitlines()
            if 'Sheet="Geo"' in line and "2024 Q2" in line
        ]
        self.assertTrue(matching)
        row = matching[0]
        self.assertIn("England and Wales=9470", row)
        self.assertIn("England=7672", row)

    def test_case_04_percentage_renders_unambiguously(self) -> None:
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws.title = "Rates"
            ws["A1"] = "Metric"
            ws["B1"] = "Share"
            ws["A2"] = "ASHP"
            ws["B2"] = 0.48
            ws["B2"].number_format = "0%"

        result = extract_xlsx_text(_xlsx_bytes(build))
        self.assertIn("48%", result.text)
        self.assertNotIn("Share=0.48", result.text)

    def test_case_05_date_year_quarter_preserved(self) -> None:
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws.title = "Timeline"
            ws["A1"] = "Installation quarter"
            ws["B1"] = "Year"
            ws["C1"] = "Event date"
            ws["A2"] = "2024 Q2: April to June"
            ws["B2"] = 2024
            ws["C2"] = date(2024, 4, 1)
            ws["C2"].number_format = "YYYY-MM-DD"

        result = extract_xlsx_text(_xlsx_bytes(build))
        self.assertIn("2024 Q2: April to June", result.text)
        self.assertIn("2024", result.text)
        self.assertIn("2024-04-01", result.text)

    def test_case_06_blank_not_zero_explicit_na_preserved(self) -> None:
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws.title = "Sparse"
            ws["A1"] = "Region"
            ws["B1"] = "Count"
            ws["A2"] = "North"
            ws["B2"] = None
            ws["A3"] = "Wales"
            ws["B3"] = "NA"
            ws["A4"] = "London"
            ws["B4"] = 0

        result = extract_xlsx_text(_xlsx_bytes(build))
        north_lines = [
            line
            for line in result.text.splitlines()
            if 'Sheet="Sparse"' in line and "North" in line
        ]
        self.assertTrue(north_lines)
        self.assertNotIn("Count=0", north_lines[0])
        self.assertNotIn("=0", north_lines[0].split("North", 1)[-1])
        wales = [
            line
            for line in result.text.splitlines()
            if 'Sheet="Sparse"' in line and "Wales" in line
        ][0]
        self.assertIn("Count=NA", wales)
        london = [
            line
            for line in result.text.splitlines()
            if 'Sheet="Sparse"' in line and "London" in line
        ][0]
        self.assertIn("Count=0", london)

    def test_case_07_merged_header_context_retained(self) -> None:
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws.title = "Merged"
            ws["A1"] = "Period"
            ws["B1"] = "United Kingdom"
            ws.merge_cells("B1:C1")
            ws["A2"] = "Label"
            ws["B2"] = "Total"
            ws["C2"] = "Of which ASHP"
            ws["A3"] = "2024"
            ws["B3"] = 100
            ws["C3"] = 90

        result = extract_xlsx_text(_xlsx_bytes(build))
        row = [
            line
            for line in result.text.splitlines()
            if 'Sheet="Merged"' in line and "2024" in line
        ][0]
        self.assertIn("United Kingdom", row)
        self.assertIn("100", row)
        self.assertIn("90", row)

    def test_case_08_cached_formula_value_without_execution(self) -> None:
        # Minimal OOXML with formula + cached value (no Excel recalculation).
        payload = _formula_cached_xlsx_bytes()
        result = extract_xlsx_text(payload)
        self.assertTrue(result.ok)
        self.assertIn("42", result.text)

    def test_case_09_cover_notes_do_not_corrupt_table(self) -> None:
        def build(wb: Workbook) -> None:
            cover = wb.active
            cover.title = "Cover"
            cover["A1"] = "Official statistics cover page"
            notes = wb.create_sheet("Notes")
            notes["A1"] = "note 1"
            notes["B1"] = "Coverage caveat"
            table = wb.create_sheet("Table 1.3")
            table["A1"] = "Installation quarter"
            table["B1"] = "England and Wales"
            table["A2"] = "2024 Q2"
            table["B2"] = 9470

        result = extract_xlsx_text(_xlsx_bytes(build))
        self.assertIn("[Sheet: Cover]", result.text)
        self.assertIn("[Sheet: Notes]", result.text)
        self.assertIn("[Sheet: Table 1.3]", result.text)
        table_rows = [
            line
            for line in result.text.splitlines()
            if 'Sheet="Table 1.3"' in line and "2024 Q2" in line
        ]
        self.assertTrue(table_rows)
        self.assertIn("England and Wales=9470", table_rows[0])

    def test_case_10_large_sheet_truncation_diagnostics(self) -> None:
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws.title = "Big"
            ws["A1"] = "Row"
            ws["B1"] = "Value"
            for idx in range(2, 80):
                ws[f"A{idx}"] = f"r{idx}"
                ws[f"B{idx}"] = idx

        result = extract_xlsx_text(
            _xlsx_bytes(build),
            limits=XlsxExtractionLimits(max_rows_per_sheet=10),
        )
        self.assertTrue(result.ok)
        self.assertTrue(result.metadata.get("workbook_truncated"))
        self.assertGreater(result.metadata.get("rows_skipped_for_bounds", 0), 0)

    def test_case_11_malformed_xlsx_controlled_failure(self) -> None:
        result = extract_xlsx_text(b"PK\x03\x04not-a-real-xlsx")
        self.assertEqual(result.error, "workbook_parse_failed")
        self.assertFalse(result.text)

    def test_case_12_non_xlsx_zip_controlled_failure(self) -> None:
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w") as archive:
            archive.writestr("readme.txt", "hello")
        result = extract_xlsx_text(buffer.getvalue())
        self.assertEqual(result.error, "workbook_parse_failed")

    @patch("infrastructure.retrieval.redirect_fetcher.validate_fetch_url")
    def test_case_13_xls_remains_explicitly_unsupported(self, _validate) -> None:
        client = Mock()
        client.get.return_value = _mock_response(
            url="https://example.com/stats.xls",
            content_type="application/vnd.ms-excel",
            content=b"\xd0\xcf\x11\xe0",  # OLE header-ish
        )
        source = HttpSourceRetriever(http_client=client).retrieve(
            SourceCandidate(
                provider="test",
                url="https://example.com/stats.xls",
                title="Old XLS",
                snippet="",
                query_id="sq-1",
                rank=1,
            ),
        )
        self.assertEqual(source.retrieval_status, RetrievalStatus.UNSUPPORTED)
        self.assertEqual(source.metadata.get("reason"), "unsupported_spreadsheet_format")
        self.assertEqual(
            source.metadata.get("failure_category"),
            "unsupported_content_type",
        )

    def test_case_14_grounded_xlsx_excerpt_passes(self) -> None:
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws.title = "Table 1.3"
            ws["A1"] = "Installation quarter"
            ws["B1"] = "England and Wales"
            ws["A2"] = "2024 Q2"
            ws["B2"] = 9470

        text = extract_xlsx_text(_xlsx_bytes(build)).text
        excerpt = [
            line
            for line in text.splitlines()
            if 'Sheet="Table 1.3"' in line and "England and Wales=9470" in line
        ][0]
        locator = verify_grounding(source_text=text, excerpt=excerpt)
        self.assertGreaterEqual(locator.normalized_start, 0)

    def test_case_15_fabricated_excerpt_fails_grounding(self) -> None:
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws.title = "Table 1.3"
            ws["A1"] = "Installation quarter"
            ws["B1"] = "England and Wales"
            ws["A2"] = "2024 Q2"
            ws["B2"] = 9470

        text = extract_xlsx_text(_xlsx_bytes(build)).text
        from application.evidence.exceptions import UngroundedEvidenceError

        with self.assertRaises(UngroundedEvidenceError):
            verify_grounding(
                source_text=text,
                excerpt='Sheet="Table 1.3" | Installation quarter=2099 Q9 | England and Wales=999999',
            )

    def test_case_16_quantitative_row_retains_metric_geo_period(self) -> None:
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws.title = "Table 1.3"
            ws["A1"] = "Installation quarter"
            ws["B1"] = "England and Wales"
            ws["A2"] = "2024 Q2"
            ws["B2"] = 43500

        text = extract_xlsx_text(_xlsx_bytes(build)).text
        row = [
            line
            for line in text.splitlines()
            if 'Sheet="Table 1.3"' in line and "43500" in line
        ][0]
        self.assertIn("Installation quarter=2024 Q2", row)
        self.assertIn("England and Wales=43500", row)

    def test_case_17_desnz_real_workbook_replay(self) -> None:
        if not DESNZ_PATH.is_file():
            self.skipTest(f"DESNZ fixture missing at {DESNZ_PATH}")
        data = DESNZ_PATH.read_bytes()
        self.assertEqual(hashlib.sha256(data).hexdigest(), DESNZ_SHA256)
        started = time.perf_counter()
        result = extract_xlsx_text(data, content_type=XLSX_CONTENT_TYPE)
        elapsed = time.perf_counter() - started
        self.assertTrue(result.ok, msg=str(result.metadata))
        self.assertIsNone(result.error)
        self.assertIn("[Sheet: Table 1.3]", result.text)
        self.assertIn("England and Wales", result.text)
        self.assertTrue(
            any(token in result.text for token in ("2022 Q", "2023 Q", "2024 Q", "2025 Q")),
        )
        self.assertTrue(
            "Methodology" in result.text or "methodology" in result.text.lower(),
        )
        self.assertIn("Note", result.text)
        self.assertLess(elapsed, 5.0)
        # HTTP integration path
        client = Mock()
        client.get.return_value = _mock_response(
            url=(
                "https://assets.publishing.service.gov.uk/media/68b5c751536d629f9c82a983/"
                "Heat_pump_deployment_quarterly_statistics_United_Kingdom_2025_Q2.xlsx"
            ),
            content_type=XLSX_CONTENT_TYPE,
            content=data,
        )
        with patch(
            "infrastructure.retrieval.redirect_fetcher.validate_fetch_url",
        ):
            source = HttpSourceRetriever(http_client=client).retrieve(
                SourceCandidate(
                    provider="test",
                    url=(
                        "https://assets.publishing.service.gov.uk/media/68b5c751536d629f9c82a983/"
                        "Heat_pump_deployment_quarterly_statistics_United_Kingdom_2025_Q2.xlsx"
                    ),
                    title="DESNZ",
                    snippet="",
                    query_id="sq-1",
                    rank=1,
                ),
            )
        self.assertIn(
            source.retrieval_status,
            {RetrievalStatus.ACQUIRED, RetrievalStatus.TRUNCATED},
        )
        self.assertTrue(source.content_text)
        self.assertNotEqual(
            source.metadata.get("failure_category"),
            "unsupported_content_type",
        )
        self.assertEqual(source.metadata.get("parser"), "xlsx")

    @patch("infrastructure.retrieval.network_safety._default_resolve_host_addresses")
    def test_case_18_html_path_unchanged(self, resolve_addresses) -> None:
        resolve_addresses.return_value = ["93.184.216.34"]
        client = Mock()
        client.get.return_value = _mock_response(
            url="https://example.com/report",
            content_type="text/html; charset=utf-8",
            content=b"<html><body><p>Hello world</p></body></html>",
        )
        source = HttpSourceRetriever(http_client=client).retrieve(
            SourceCandidate(
                provider="test",
                url="https://example.com/report",
                title="Report",
                snippet="",
                query_id="sq-1",
                rank=1,
            ),
        )
        self.assertEqual(source.retrieval_status, RetrievalStatus.ACQUIRED)
        self.assertIn("Hello world", source.content_text)

    @patch("infrastructure.retrieval.redirect_fetcher.validate_fetch_url")
    def test_case_19_pdf_path_unchanged(self, _validate) -> None:
        client = Mock()
        client.get.return_value = _mock_response(
            url="https://example.com/report.pdf",
            content_type="application/pdf",
            content=b"%PDF-1.4",
        )
        source = HttpSourceRetriever(http_client=client).retrieve(
            SourceCandidate(
                provider="test",
                url="https://example.com/report.pdf",
                title="PDF",
                snippet="",
                query_id="sq-1",
                rank=1,
            ),
        )
        self.assertEqual(source.retrieval_status, RetrievalStatus.UNSUPPORTED)
        self.assertEqual(
            source.metadata.get("reason"),
            "PDF retrieval deferred in DR-03 v1",
        )

    def test_case_20_p1_12_source_ranking_module_untouched_contract(self) -> None:
        from application.sources.deterministic_source_relevance import (
            evaluate_candidate,
            selection_sort_key,
        )

        self.assertTrue(callable(evaluate_candidate))
        self.assertTrue(callable(selection_sort_key))

    def test_case_21_no_llm_call_used_by_xlsx_parser(self) -> None:
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws.title = "Only"
            ws["A1"] = "x"
            ws["B1"] = 1

        with patch("openai.OpenAI") as openai_cls:
            result = extract_xlsx_text(_xlsx_bytes(build))
            self.assertTrue(result.ok)
            openai_cls.assert_not_called()

    def test_case_22_hidden_sheets_skipped_and_diagnosed(self) -> None:
        def build(wb: Workbook) -> None:
            visible = wb.active
            visible.title = "Visible"
            visible["A1"] = "keep"
            hidden = wb.create_sheet("HiddenAdmin")
            hidden["A1"] = "secret"
            hidden.sheet_state = "hidden"

        result = extract_xlsx_text(_xlsx_bytes(build))
        self.assertIn("[Sheet: Visible]", result.text)
        self.assertNotIn("HiddenAdmin", result.text)
        self.assertEqual(result.metadata.get("skipped_hidden_sheets"), 1)
        self.assertIn("HiddenAdmin", result.metadata.get("skipped_sheet_names", []))

    def test_case_23_workbook_over_max_sheets_bounded(self) -> None:
        def build(wb: Workbook) -> None:
            wb.active.title = "S1"
            wb.active["A1"] = "one"
            for idx in range(2, 6):
                sheet = wb.create_sheet(f"S{idx}")
                sheet["A1"] = f"v{idx}"

        result = extract_xlsx_text(
            _xlsx_bytes(build),
            limits=XlsxExtractionLimits(max_sheets=2),
        )
        self.assertTrue(result.ok)
        self.assertTrue(result.metadata.get("workbook_truncated"))
        self.assertEqual(result.metadata.get("processed_sheets"), 2)
        self.assertGreaterEqual(result.metadata.get("sheets_skipped_for_bounds", 0), 1)
        self.assertIn("[Sheet: S1]", result.text)
        self.assertIn("[Sheet: S2]", result.text)
        self.assertNotIn("[Sheet: S5]", result.text)

    def test_case_24_workbook_over_max_rendered_chars_bounded(self) -> None:
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws.title = "Wide"
            ws["A1"] = "Label"
            ws["B1"] = "Value"
            for idx in range(2, 40):
                ws[f"A{idx}"] = f"row-{idx}-" + ("x" * 40)
                ws[f"B{idx}"] = idx

        result = extract_xlsx_text(
            _xlsx_bytes(build),
            limits=XlsxExtractionLimits(max_rendered_chars=250),
        )
        self.assertTrue(result.text)
        self.assertTrue(result.metadata.get("workbook_truncated"))
        self.assertLessEqual(len(result.text), 250)

    def test_evidence_pipeline_can_consume_rendered_text_without_new_stage(self) -> None:
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws.title = "Table 1.3"
            ws["A1"] = "Installation quarter"
            ws["B1"] = "England and Wales"
            ws["A2"] = "2024 Q2"
            ws["B2"] = 9470

        text = extract_xlsx_text(_xlsx_bytes(build)).text
        from domain.sources.source import Source
        from application.evidence.content_chunking import split_normalized_source_content

        source = Source(
            id="src-1",
            project_id="p1",
            url="https://example.com/t.xlsx",
            canonical_url="https://example.com/t.xlsx",
            title="t",
            retrieved_at="2026-01-01T00:00:00Z",
            content_type=XLSX_CONTENT_TYPE,
            retrieval_status=RetrievalStatus.ACQUIRED,
            content_text=text,
            content_checksum=hashlib.sha256(text.encode("utf-8")).hexdigest(),
        )
        chunks = split_normalized_source_content(
            source.content_text,
            chunk_chars=2000,
            overlap_chars=100,
        )
        self.assertTrue(chunks)
        self.assertTrue(any("England and Wales=9470" in chunk.text for chunk in chunks))


def _formula_cached_xlsx_bytes() -> bytes:
    """Build a tiny XLSX containing a formula cell with cached value 42."""
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(
            "[Content_Types].xml",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>
</Types>
""",
        )
        archive.writestr(
            "_rels/.rels",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>
""",
        )
        archive.writestr(
            "xl/workbook.xml",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
    <sheet name="Calc" sheetId="1" r:id="rId1"/>
  </sheets>
</workbook>
""",
        )
        archive.writestr(
            "xl/_rels/workbook.xml.rels",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
</Relationships>
""",
        )
        archive.writestr(
            "xl/sharedStrings.xml",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="2" uniqueCount="2">
  <si><t>Metric</t></si>
  <si><t>Total</t></si>
</sst>
""",
        )
        archive.writestr(
            "xl/worksheets/sheet1.xml",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <sheetData>
    <row r="1">
      <c r="A1" t="s"><v>0</v></c>
      <c r="B1" t="s"><v>1</v></c>
    </row>
    <row r="2">
      <c r="A2" t="s"><v>0</v></c>
      <c r="B2"><f>40+2</f><v>42</v></c>
    </row>
  </sheetData>
</worksheet>
""",
        )
    return buffer.getvalue()


if __name__ == "__main__":
    unittest.main()

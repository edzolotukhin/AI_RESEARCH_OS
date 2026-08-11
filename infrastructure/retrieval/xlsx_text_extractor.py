"""Deterministic bounded XLSX → text extraction for Source.content_text.

P1-15.1: no LLM, no formula execution, no macros/external links.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime, time
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from openpyxl.styles.numbers import is_date_format


XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

UNSUPPORTED_SPREADSHEET_EXTENSIONS = (".xls", ".xlsb", ".ods")

MAX_SHEETS = 20
MAX_ROWS_PER_SHEET = 500
MAX_COLUMNS_PER_SHEET = 40
MAX_RENDERED_CHARS = 200_000

_WHITESPACE_RE = re.compile(r"\s+")
_EXPLICIT_NA_MARKERS = frozenset(
    {
        "na",
        "n/a",
        "n.a.",
        "not available",
        "not applicable",
        "suppressed",
        "c",
        "..",
        "-",
    }
)


@dataclass(frozen=True)
class XlsxExtractionLimits:
    max_sheets: int = MAX_SHEETS
    max_rows_per_sheet: int = MAX_ROWS_PER_SHEET
    max_columns_per_sheet: int = MAX_COLUMNS_PER_SHEET
    max_rendered_chars: int = MAX_RENDERED_CHARS


@dataclass
class XlsxExtractionResult:
    text: str = ""
    error: str | None = None
    metadata: dict[str, Any] = field(default_factory=dict)

    @property
    def ok(self) -> bool:
        return self.error is None and bool(self.text.strip())


def is_xlsx_content_type(content_type: str) -> bool:
    return (content_type or "").split(";")[0].strip().lower().startswith(
        XLSX_CONTENT_TYPE
    )


def looks_like_xlsx_url(url: str) -> bool:
    path = (url or "").split("?", 1)[0].lower()
    return path.endswith(".xlsx")


def is_unsupported_spreadsheet_url(url: str) -> bool:
    path = (url or "").split("?", 1)[0].lower()
    if path.endswith(".xlsx"):
        return False
    return any(path.endswith(ext) for ext in UNSUPPORTED_SPREADSHEET_EXTENSIONS)


def extract_xlsx_text(
    data: bytes,
    *,
    limits: XlsxExtractionLimits | None = None,
    content_type: str = "",
) -> XlsxExtractionResult:
    """Parse XLSX bytes into bounded deterministic content_text + diagnostics."""
    bounds = limits or XlsxExtractionLimits()
    meta: dict[str, Any] = {
        "parser": "xlsx",
        "content_type": content_type or XLSX_CONTENT_TYPE,
        "sheet_count": 0,
        "visible_sheets": 0,
        "processed_sheets": 0,
        "skipped_hidden_sheets": 0,
        "skipped_sheet_names": [],
        "rows_processed": 0,
        "cells_processed": 0,
        "rendered_chars": 0,
        "workbook_truncated": False,
        "sheets_skipped_for_bounds": 0,
        "rows_skipped_for_bounds": 0,
    }

    if not data:
        meta["parse_error"] = "empty_body"
        return XlsxExtractionResult(error="workbook_parse_failed", metadata=meta)

    if data[:2] != b"PK":
        meta["parse_error"] = "not_zip_container"
        return XlsxExtractionResult(error="workbook_parse_failed", metadata=meta)

    try:
        workbook = load_workbook(
            filename=io.BytesIO(data),
            read_only=False,
            data_only=True,
            keep_vba=False,
        )
    except Exception as exc:  # noqa: BLE001
        message = str(exc).lower()
        if "encrypt" in message or "password" in message:
            meta["parse_error"] = "encrypted_workbook"
            return XlsxExtractionResult(error="encrypted_workbook", metadata=meta)
        meta["parse_error"] = f"workbook_parse_failed:{type(exc).__name__}"
        return XlsxExtractionResult(error="workbook_parse_failed", metadata=meta)

    try:
        return _render_workbook(workbook, bounds=bounds, meta=meta)
    except Exception as exc:  # noqa: BLE001
        meta["parse_error"] = f"workbook_render_failed:{type(exc).__name__}"
        return XlsxExtractionResult(error="workbook_parse_failed", metadata=meta)
    finally:
        try:
            workbook.close()
        except Exception:  # noqa: BLE001
            pass


def _render_workbook(
    workbook,
    *,
    bounds: XlsxExtractionLimits,
    meta: dict[str, Any],
) -> XlsxExtractionResult:
    sheet_names = list(workbook.sheetnames)
    meta["sheet_count"] = len(sheet_names)

    visible_names: list[str] = []
    for name in sheet_names:
        sheet = workbook[name]
        state = getattr(sheet, "sheet_state", "visible") or "visible"
        if state != "visible":
            meta["skipped_hidden_sheets"] += 1
            meta["skipped_sheet_names"].append(name)
            continue
        visible_names.append(name)

    meta["visible_sheets"] = len(visible_names)

    if len(visible_names) > bounds.max_sheets:
        meta["workbook_truncated"] = True
        meta["sheets_skipped_for_bounds"] = len(visible_names) - bounds.max_sheets
        visible_names = visible_names[: bounds.max_sheets]

    chunks: list[str] = []
    rendered_chars = 0

    for name in visible_names:
        if rendered_chars >= bounds.max_rendered_chars:
            meta["workbook_truncated"] = True
            break

        sheet_text, sheet_stats = _render_sheet(
            workbook[name],
            sheet_name=name,
            bounds=bounds,
            remaining_chars=bounds.max_rendered_chars - rendered_chars,
        )
        meta["processed_sheets"] += 1
        meta["rows_processed"] += sheet_stats["rows_processed"]
        meta["cells_processed"] += sheet_stats["cells_processed"]
        meta["rows_skipped_for_bounds"] += sheet_stats["rows_skipped_for_bounds"]
        if sheet_stats["truncated"]:
            meta["workbook_truncated"] = True

        if not sheet_text:
            continue

        if chunks:
            piece = "\n\n" + sheet_text
        else:
            piece = sheet_text

        if rendered_chars + len(piece) > bounds.max_rendered_chars:
            room = bounds.max_rendered_chars - rendered_chars
            if room <= 0:
                meta["workbook_truncated"] = True
                break
            piece = piece[:room]
            meta["workbook_truncated"] = True
            chunks.append(piece)
            rendered_chars += len(piece)
            break

        chunks.append(piece)
        rendered_chars += len(piece)

    text = "".join(chunks).strip()
    meta["rendered_chars"] = len(text)

    if not text:
        meta["parse_error"] = "no_renderable_cells"
        return XlsxExtractionResult(error="no_renderable_cells", metadata=meta)

    return XlsxExtractionResult(text=text, metadata=meta)


def _render_sheet(
    sheet,
    *,
    sheet_name: str,
    bounds: XlsxExtractionLimits,
    remaining_chars: int,
) -> tuple[str, dict[str, int | bool]]:
    stats: dict[str, int | bool] = {
        "rows_processed": 0,
        "cells_processed": 0,
        "rows_skipped_for_bounds": 0,
        "truncated": False,
    }

    merge_map = _build_merge_map(sheet)
    grid, max_row, max_col, row_trunc = _materialize_grid(
        sheet,
        merge_map=merge_map,
        max_rows=bounds.max_rows_per_sheet,
        max_cols=bounds.max_columns_per_sheet,
    )
    stats["rows_skipped_for_bounds"] = row_trunc
    if row_trunc:
        stats["truncated"] = True

    if max_row == 0 or max_col == 0:
        return "", stats

    lines: list[str] = [f'[Sheet: {sheet_name}]']
    header_end, headers = _detect_headers(grid, max_row=max_row, max_col=max_col)
    table_mode = _is_table_mode(grid, headers, header_end, max_row=max_row, max_col=max_col)

    if table_mode and any(headers):
        header_line = " | ".join(
            h for h in headers if h
        )
        if header_line:
            lines.append(f"headers: {header_line}")

    for row_idx in range(1, max_row + 1):
        row_vals = [
            _cell_text(grid.get(row_idx, {}).get(col))
            for col in range(1, max_col + 1)
        ]
        nonempty = [(col, value) for col, value in enumerate(row_vals, start=1) if value]
        if not nonempty:
            continue

        stats["rows_processed"] += 1
        stats["cells_processed"] += len(nonempty)

        if table_mode and row_idx > header_end and any(headers):
            parts = [f'Sheet="{sheet_name}"']
            for col, value in nonempty:
                label = headers[col - 1] if col - 1 < len(headers) else ""
                if not label:
                    label = f"col_{col}"
                parts.append(f"{label}={value}")
            line = " | ".join(parts)
        elif table_mode and row_idx <= header_end:
            # Header / title rows already represented; still keep title-like prose.
            if row_idx < header_end or not any(headers):
                joined = " | ".join(value for _, value in nonempty)
                line = joined
            else:
                continue
        else:
            joined = " | ".join(value for _, value in nonempty)
            line = joined

        lines.append(line)
        probe = "\n".join(lines)
        if len(probe) > remaining_chars:
            stats["truncated"] = True
            # Drop the overflowing line and stop.
            lines.pop()
            break

    text = "\n".join(lines).strip()
    if text == f'[Sheet: {sheet_name}]':
        return "", stats
    return text, stats


def _build_merge_map(sheet) -> dict[tuple[int, int], tuple[int, int]]:
    mapping: dict[tuple[int, int], tuple[int, int]] = {}
    merged_ranges = getattr(sheet, "merged_cells", None)
    if merged_ranges is None:
        return mapping
    ranges = getattr(merged_ranges, "ranges", merged_ranges)
    for merged in list(ranges):
        min_row = int(merged.min_row)
        min_col = int(merged.min_col)
        max_row = int(merged.max_row)
        max_col = int(merged.max_col)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if (row, col) != (min_row, min_col):
                    mapping[(row, col)] = (min_row, min_col)
    return mapping


def _materialize_grid(
    sheet,
    *,
    merge_map: dict[tuple[int, int], tuple[int, int]],
    max_rows: int,
    max_cols: int,
) -> tuple[dict[int, dict[int, Any]], int, int, int]:
    """Return grid[row][col] = openpyxl cell (or synthetic), plus dimensions."""
    grid: dict[int, dict[int, Any]] = {}
    max_row = 0
    max_col = 0
    rows_seen = 0
    truncated_rows = 0

    for row in sheet.iter_rows(
        min_row=1,
        max_row=None,
        max_col=max_cols,
        values_only=False,
    ):
        rows_seen += 1
        if rows_seen > max_rows:
            truncated_rows += 1
            continue

        row_idx = row[0].row if row else rows_seen
        grid[row_idx] = {}
        for cell in row:
            col_idx = cell.column
            if col_idx > max_cols:
                continue
            source = cell
            origin = merge_map.get((row_idx, col_idx))
            if origin is not None:
                origin_row, origin_col = origin
                if origin_row in grid and origin_col in grid[origin_row]:
                    source = grid[origin_row][origin_col]
                else:
                    source = sheet.cell(row=origin_row, column=origin_col)
            if source.value is None:
                continue
            grid[row_idx][col_idx] = source
            max_row = max(max_row, row_idx)
            max_col = max(max_col, col_idx)

        # Stop scanning once bounds exceeded to avoid walking huge sheets.
        if truncated_rows:
            # Count remaining via sheet.max_row when available, then stop.
            sheet_max_row = getattr(sheet, "max_row", None) or 0
            if isinstance(sheet_max_row, int) and sheet_max_row > max_rows:
                truncated_rows = max(truncated_rows, sheet_max_row - max_rows)
            break

    # Fill merge-propagated empties for rows already captured.
    for (row_idx, col_idx), (origin_row, origin_col) in merge_map.items():
        if row_idx > max_rows or col_idx > max_cols:
            continue
        if row_idx not in grid:
            continue
        if col_idx in grid[row_idx]:
            continue
        origin_cell = None
        if origin_row in grid and origin_col in grid[origin_row]:
            origin_cell = grid[origin_row][origin_col]
        else:
            try:
                origin_cell = sheet.cell(row=origin_row, column=origin_col)
            except Exception:  # noqa: BLE001
                origin_cell = None
        if origin_cell is None or origin_cell.value is None:
            continue
        grid[row_idx][col_idx] = origin_cell
        max_row = max(max_row, row_idx)
        max_col = max(max_col, col_idx)

    if truncated_rows == 0:
        sheet_max_row = getattr(sheet, "max_row", None) or 0
        if isinstance(sheet_max_row, int) and sheet_max_row > max_rows:
            truncated_rows = sheet_max_row - max_rows

    return grid, max_row, min(max_col, max_cols), truncated_rows


def _cell_text(cell: Any | None) -> str:
    if cell is None:
        return ""
    return format_cell_display(cell)


def format_cell_display(cell: Any) -> str:
    """Render a cell value with number-format-aware semantics."""
    value = getattr(cell, "value", cell)
    if value is None:
        return ""

    number_format = str(getattr(cell, "number_format", "") or "")

    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"

    if isinstance(value, datetime):
        if value.time() == time(0, 0):
            return value.date().isoformat()
        return value.isoformat(sep=" ", timespec="seconds")

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, time):
        return value.isoformat(timespec="seconds")

    if isinstance(value, (int, float)):
        return _format_number(value, number_format)

    text = _normalize_whitespace(str(value))
    return text


def _format_number(value: int | float, number_format: str) -> str:
    fmt = number_format or "General"
    fmt_lower = fmt.lower()

    if "%" in fmt:
        # Excel stores percentages as fractions (0.48 → 48%).
        pct = float(value) * 100.0
        if abs(pct - round(pct)) < 1e-9:
            return f"{int(round(pct))}%"
        return f"{_trim_float(pct)}%"

    if is_date_format(fmt) and not isinstance(value, (datetime, date)):
        try:
            converted = from_excel(value)
            if isinstance(converted, datetime):
                if converted.time() == time(0, 0):
                    return converted.date().isoformat()
                return converted.isoformat(sep=" ", timespec="seconds")
            if isinstance(converted, date):
                return converted.isoformat()
        except Exception:  # noqa: BLE001
            pass

    if isinstance(value, float):
        if abs(value - round(value)) < 1e-9:
            return str(int(round(value)))
        return _trim_float(value)

    return str(value)


def _trim_float(value: float) -> str:
    text = f"{value:.12g}"
    if "e" in text.lower():
        text = f"{value:.12f}".rstrip("0").rstrip(".")
    return text


def _normalize_whitespace(text: str) -> str:
    return _WHITESPACE_RE.sub(" ", text).strip()


def _detect_headers(
    grid: dict[int, dict[int, Any]],
    *,
    max_row: int,
    max_col: int,
) -> tuple[int, list[str]]:
    """Return (last_header_row_index, composed_headers_by_column)."""
    if max_row == 0:
        return 0, []

    row_profiles: list[tuple[int, int, int, int]] = []
    # (row_idx, nonempty, numeric_count, string_count)
    for row_idx in range(1, max_row + 1):
        nonempty = 0
        numeric = 0
        strings = 0
        data_like = 0
        for col in range(1, max_col + 1):
            cell = grid.get(row_idx, {}).get(col)
            text = _cell_text(cell)
            if not text:
                continue
            nonempty += 1
            value = getattr(cell, "value", None)
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                numeric += 1
                data_like += 1
            elif text.lower() in _EXPLICIT_NA_MARKERS:
                data_like += 1
            else:
                strings += 1
        row_profiles.append((row_idx, nonempty, numeric, strings, data_like))

    # Skip leading title/metadata rows with very few columns.
    first_wide = 1
    for row_idx, nonempty, _numeric, _strings, _data_like in row_profiles:
        if nonempty >= max(2, min(3, max_col)):
            first_wide = row_idx
            break

    header_rows: list[int] = []
    data_row = None
    for row_idx, nonempty, numeric, strings, data_like in row_profiles:
        if row_idx < first_wide:
            continue
        if nonempty == 0:
            continue
        # Data row: has numeric/NA values and is not string-dominated.
        if nonempty >= 2 and data_like >= 1 and data_like >= strings:
            data_row = row_idx
            break
        if nonempty >= 1 and numeric >= 1 and numeric >= strings:
            data_row = row_idx
            break
        if header_rows:
            # Extra header rows must remain wide, string-only (multi-row headers).
            if nonempty < 2 or data_like > 0:
                data_row = row_idx
                break
        header_rows.append(row_idx)
        # Cap header depth.
        if len(header_rows) >= 5:
            break

    if not header_rows:
        # Fall back: first non-empty row as headers if a later numeric row exists.
        for row_idx, nonempty, numeric, strings, data_like in row_profiles:
            if nonempty:
                header_rows = [row_idx]
                break

    if not header_rows:
        return 0, []

    # Keep only the contiguous header block immediately above data (or last strings).
    if data_row is not None:
        header_rows = [r for r in header_rows if r < data_row]
        # Prefer the last 1–3 header rows closest to data.
        header_rows = header_rows[-3:]

    composed: list[str] = []
    for col in range(1, max_col + 1):
        parts: list[str] = []
        for row_idx in header_rows:
            text = _cell_text(grid.get(row_idx, {}).get(col))
            if not text:
                continue
            if parts and parts[-1] == text:
                continue
            parts.append(text)
        composed.append(_normalize_whitespace(" / ".join(parts)))

    header_end = header_rows[-1] if header_rows else 0
    return header_end, composed


def _is_table_mode(
    grid: dict[int, dict[int, Any]],
    headers: list[str],
    header_end: int,
    *,
    max_row: int,
    max_col: int,
) -> bool:
    if max_col < 2:
        return False
    if sum(1 for h in headers if h) < 2:
        return False
    for row_idx in range(header_end + 1, max_row + 1):
        numeric = 0
        nonempty = 0
        for col in range(1, max_col + 1):
            cell = grid.get(row_idx, {}).get(col)
            text = _cell_text(cell)
            if not text:
                continue
            nonempty += 1
            value = getattr(cell, "value", None)
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                numeric += 1
            elif text.lower() in _EXPLICIT_NA_MARKERS:
                numeric += 1
        if nonempty >= 2 and numeric >= 1:
            return True
    return False

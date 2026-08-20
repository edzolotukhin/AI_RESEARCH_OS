from __future__ import annotations

import io
from collections import Counter
from typing import Any

import openpyxl

from application.ports.quantitative_dataset_ports import ParsedDataset, ParsedVariable
from domain.quantitative.dataset import DatasetFormat


class XlsxOpenpyxlAdapter:
    format = DatasetFormat.XLSX

    def parse(
        self,
        data: bytes,
        *,
        filename: str,
        data_sheet: str | None = None,
    ) -> ParsedDataset:
        formulas = openpyxl.load_workbook(
            io.BytesIO(data), read_only=False, data_only=False, keep_vba=False
        )
        values = openpyxl.load_workbook(
            io.BytesIO(data), read_only=False, data_only=True, keep_vba=False
        )
        try:
            sheet_name = _resolve_sheet(formulas, data_sheet)
            formula_sheet = formulas[sheet_name]
            value_sheet = values[sheet_name]
            warnings: list[str] = []
            if formula_sheet.sheet_state != "visible":
                warnings.append("selected_data_sheet_hidden")
            if any(formula_sheet.row_dimensions[index].hidden for index in formula_sheet.row_dimensions):
                warnings.append("hidden_rows_present")
            if any(formula_sheet.column_dimensions[key].hidden for key in formula_sheet.column_dimensions):
                warnings.append("hidden_columns_present")

            raw_rows = list(value_sheet.iter_rows(values_only=True))
            formula_rows = list(formula_sheet.iter_rows(values_only=False))
            while raw_rows and all(value is None for value in raw_rows[-1]):
                raw_rows.pop()
                formula_rows.pop()
            if not raw_rows:
                raise ValueError("XLSX data sheet is empty")
            headers = tuple("" if value is None else str(value).strip() for value in raw_rows[0])
            if any(not item for item in headers):
                raise ValueError("XLSX headers must be non-empty")
            normalized = [item.casefold() for item in headers]
            if len(set(normalized)) != len(normalized):
                raise ValueError("XLSX headers must be unique")

            rows: list[tuple[Any, ...]] = []
            for row_number, values_row in enumerate(raw_rows[1:], start=2):
                trimmed = tuple(values_row[: len(headers)])
                if all(value is None for value in trimmed):
                    warnings.append(f"internal_empty_row:{row_number}")
                    rows.append(trimmed)
                    continue
                formula_row = formula_rows[row_number - 1]
                for column_index, cell in enumerate(formula_row[: len(headers)]):
                    if cell.data_type == "f":
                        warnings.append(
                            f"formula_stored_value_only:{row_number}:{column_index + 1}"
                        )
                rows.append(trimmed)

            variables = tuple(
                ParsedVariable(
                    name=header,
                    storage_type=_storage_type([row[index] for row in rows]),
                    metadata={
                        "mixed_types": _mixed_types([row[index] for row in rows]),
                    },
                )
                for index, header in enumerate(headers)
            )
            return ParsedDataset(
                format=DatasetFormat.XLSX,
                variables=variables,
                rows=tuple(rows),
                parser_name="openpyxl",
                parser_version=openpyxl.__version__,
                warnings=tuple(dict.fromkeys(warnings)),
            )
        finally:
            formulas.close()
            values.close()


def _resolve_sheet(workbook: Any, requested: str | None) -> str:
    if requested is not None:
        if requested not in workbook.sheetnames:
            raise ValueError(f"XLSX data sheet not found: {requested}")
        return requested
    visible = [item.title for item in workbook.worksheets if item.sheet_state == "visible"]
    if len(visible) != 1:
        raise ValueError("XLSX data sheet must be explicitly selected")
    return visible[0]


def _value_kind(value: Any) -> str:
    if value is None:
        return "missing"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "numeric"
    return "string"


def _mixed_types(values: list[Any]) -> bool:
    kinds = {_value_kind(value) for value in values} - {"missing"}
    return len(kinds) > 1


def _storage_type(values: list[Any]) -> str:
    counts = Counter(_value_kind(value) for value in values if value is not None)
    return counts.most_common(1)[0][0] if counts else "unknown"
